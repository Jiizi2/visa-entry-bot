from __future__ import annotations

import re
from datetime import date, timedelta

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - depends on local environment
    load_workbook = None

COLUMN_MAP = {
    "NO": "rowNumber",
    "GENDER": "title",
    "NAMA": "fullName",
    "SEX": "gender",
    "POB": "placeOfBirth",
    "DOB": "dob",
    "ISSUING OFFICE": "issuingOffice",
    "PASSPORT": "passportNumber",
    "NATIONALITY": "nationality",
    "DOI": "issueDate",
    "DOE": "expiryDate",
    "KET": "notes",
}
PASSENGER_COLUMN_MAP = {
    1: "title",
    2: "fullName",
    3: "gender",
    4: "placeOfBirth",
    5: "dob",
    6: "passportNumber",
    7: "issuingOffice",
    8: "issueDate",
    9: "expiryDate",
}
MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


def load_reference_workbook(path: str) -> list[dict[str, str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed.")
    rows = _load_workbook_rows(path)
    header_index = _find_header_row(rows)
    if header_index >= 0:
        return _load_standard_rows(rows, header_index)
    header_index = _find_passenger_manifest_row(rows)
    if header_index >= 0:
        return _load_passenger_manifest_rows(rows, header_index)
    raise ValueError("Reference header row not found.")


def _load_workbook_rows(path: str) -> list[list[object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _load_standard_rows(rows_data: list[list[object]], header_index: int) -> list[dict[str, str]]:
    header = [_normalize_header(value) for value in rows_data[header_index]]
    rows: list[dict[str, str]] = []
    for raw_row in rows_data[header_index + 1 :]:
        record = _map_row(header, raw_row)
        if not record.get("fullName") and not record.get("passportNumber"):
            continue
        rows.append(record)
    return rows


def _load_passenger_manifest_rows(rows_data: list[list[object]], header_index: int) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for raw_row in rows_data[header_index + 1 :]:
        record = _map_passenger_row(raw_row)
        if record:
            rows.append(record)
    return rows

def split_full_name(value: str) -> tuple[str, str]:
    tokens = _normalize_text(value).split()
    if not tokens:
        return "", ""
    if len(tokens) == 1:
        return tokens[0], tokens[0]
    return " ".join(tokens[:-1]), tokens[-1]


def normalize_reference_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", _normalize_text(value))


def _find_header_row(rows_data: list[list[object]]) -> int:
    for index, row in enumerate(rows_data):
        values = {_normalize_header(value) for value in row}
        if "NAMA" in values and "PASSPORT" in values and "DOB" in values:
            return index
    return -1


def _find_passenger_manifest_row(rows_data: list[list[object]]) -> int:
    for index, row in enumerate(rows_data):
        values = {_normalize_header(value) for value in row}
        if "PASSENGER NAME" in values and "NO. PASPORT" in values and "ISSUING OFFICE" in values:
            return index
    return -1


def _map_row(header: list[str], values: list[object]) -> ParsedPassportData:
    mapped: dict[str, str] = {}
    for key, value in zip(header, values):
        target_key = COLUMN_MAP.get(key)
        if not target_key:
            continue
        mapped[target_key] = _normalize_value(target_key, value)

    first_name, family_name = split_full_name(mapped.get("fullName", ""))
    mapped["firstName"] = first_name
    mapped["familyName"] = family_name
    return mapped


def _map_passenger_row(values: list[object]) -> ParsedPassportData:
    mapped: dict[str, str] = {}
    for index, value in enumerate(values):
        target_key = PASSENGER_COLUMN_MAP.get(index)
        if not target_key:
            continue
        mapped[target_key] = _normalize_value(target_key, value)
    full_name = mapped.get("fullName", "")
    if not full_name or full_name in {"KETERANGAN", "TOTAL"}:
        return {}
    if "PENUMPANG" in full_name or "DEWASA" in full_name or "ANAK" in full_name:
        return {}
    first_name, family_name = split_full_name(full_name)
    mapped["firstName"] = first_name
    mapped["familyName"] = family_name
    return mapped


def _normalize_value(key: str, value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    if key in {"dob", "issueDate", "expiryDate"}:
        return _normalize_date(value)
    if key == "passportNumber":
        return re.sub(r"[^A-Z0-9]", "", str(value).upper())
    if key == "gender":
        gender = _normalize_text(value)
        if gender.startswith("M"):
            return "MALE"
        if gender.startswith("F"):
            return "FEMALE"
        return gender
    return _normalize_text(value)


def _normalize_date(value: object) -> str:
    if hasattr(value, "date"):
        try:
            return value.date().isoformat()
        except Exception:  # noqa: BLE001
            pass
    if isinstance(value, date):
        return value.isoformat()

    text = _normalize_text(value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    slash_match = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", text)
    if slash_match:
        day, month, year = slash_match.groups()
        try:
            return date(int(year), int(month), int(day)).isoformat()
        except ValueError:
            return ""
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        serial = int(float(text))
        if 20_000 <= serial <= 80_000:
            try:
                return (date(1899, 12, 30) + timedelta(days=serial)).isoformat()
            except ValueError:
                return ""
    match = re.search(r"\b(\d{1,2})\s+([A-Z]{3})\s+(\d{4})\b", text)
    if not match:
        return ""
    day, month_text, year = match.groups()
    month = MONTHS.get(month_text)
    if month is None:
        return ""
    try:
        return date(int(year), month, int(day)).isoformat()
    except ValueError:
        return ""


def _normalize_header(value: object) -> str:
    return _normalize_text(value)


def _normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").upper()).strip()
